import openpyxl

def get_xlsx_data(document_path, keys):
    workbook = openpyxl.load_workbook(document_path);

    sheet = workbook.active;

    data = [];

    #first_row_values = [cell.value for cell in next(sheet.iter_cols(min_row = 2, max_row = 4, values_only = True))][0];

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            print(sheet.cell(row = i, column = j).value);
        print();

    """for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row contains headers
        row_data = tuple(value for value in row)#{sheet.cell(row = 1, column = i).value: value for i, value in enumerate(row, start = 1)}
        print(row_data);
        data.append(row_data)
    """

    workbook.close()

    print(data);