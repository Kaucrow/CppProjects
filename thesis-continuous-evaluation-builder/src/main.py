import sys
import getopt
import shutil
import os
import re
import openpyxl
from openpyxl.styles import Alignment
from tqdm import tqdm

def clear():
    if os.name == 'nt':
        _ = os.system('cls')
    else:
        _ = os.system('clear')

clear()

class Globals:
    DEBUG: bool = False
    DATA_FOLDER: str = 'data/'

global globals
globals = Globals()

try:
    opts, args = getopt.getopt(sys.argv[1:], 'hd', ['help', 'debug'])
except getopt.GetoptError as err:
    print(err)
    exit(1)

for opt, arg in opts:
    if opt in ('-h', '--help'):
        print(
            '\n'.join(
                line.strip() for line in
                """
                [-h, --help]: Displays this help screen.
                [-d, --debug]: Runs the program in debug mode.
                """.split('\n')
            )
        )
        exit(0)

    if opt in ('-d', '--debug'):
        globals.DEBUG = True

keys = ['no', 'surname', 'name', 'ci', 'phone', 'email', 'thesis', 'academical_tutor', 'section', 'methodological_tutor']
records = []

wb = openpyxl.load_workbook(globals.DATA_FOLDER + 'in/architecture-thesis-records-2025A.xlsx')

sheet = wb['REGISTRO TEG I 2025A']

for row in range(4, sheet.max_row + 1):

    record = []

    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)

        cleaned_cell_text = (re.sub(r' +', ' ',
                                '\n'.join(
                                    line.strip() for line in str(cell.value).split('\n'))
                                )
                            ).strip()
        record.append(cleaned_cell_text)

    row_data = dict(zip(keys, record))

    records.append(row_data)

wb.close()

try:
    shutil.rmtree(globals.DATA_FOLDER + 'out/')
except Exception:
    pass

os.makedirs(globals.DATA_FOLDER + 'out/')

with tqdm(total=len(records), leave = True) as pbar_rec:
    for record in records:
        pbar_rec.set_description(f'Record: {record['surname'] + ' ' + record['name']}')

        wb = openpyxl.load_workbook(globals.DATA_FOLDER + 'in/continuous-evaluation-format.xlsx', data_only=True)

        method_sheet = wb['T.M.']
        academ_sheet = wb['T.A.']

        # Student
        method_cell = method_sheet.cell(row=4, column=3)
        method_cell.value = record['surname'] + ' ' + record['name']
        method_cell.alignment = Alignment(horizontal='left')

        academ_cell = academ_sheet.cell(row=4, column=3)
        academ_cell.value = record['surname'] + ' ' + record['name']
        academ_cell.alignment = Alignment(horizontal='left')

        # Thesis title
        method_cell = method_sheet.cell(row=5, column=3)
        method_cell.value = record['thesis']
        method_cell.alignment = Alignment(horizontal='left')

        academ_cell = academ_sheet.cell(row=5, column=3)
        academ_cell.value = record['thesis']
        academ_cell.alignment = Alignment(horizontal='left')

        # Section
        method_cell = method_sheet.cell(row=4, column=6)
        method_cell.value = record['section']
        method_cell.alignment = Alignment(horizontal='left')
        
        academ_cell = academ_sheet.cell(row=4, column=6)
        academ_cell.value = record['section']
        academ_cell.alignment = Alignment(horizontal='left')

        # C.I.
        method_cell = method_sheet.cell(row=4, column=8)
        method_cell.value = record['ci']
        method_cell.alignment = Alignment(horizontal='left')
        
        academ_cell = academ_sheet.cell(row=4, column=8)
        academ_cell.value = record['ci']
        academ_cell.alignment = Alignment(horizontal='left')

        # Methodological tutor 
        method_cell = method_sheet.cell(row=6, column=3)
        method_cell.value = record['methodological_tutor']
        method_cell.alignment = Alignment(horizontal='left')
        
        academ_cell = academ_sheet.cell(row=6, column=3)
        academ_cell.value = record['methodological_tutor']
        academ_cell.alignment = Alignment(horizontal='left')
        
        # Academical tutor 
        method_cell = method_sheet.cell(row=6, column=7)
        method_cell.value = record['academical_tutor']
        method_cell.alignment = Alignment(horizontal='left')
        
        academ_cell = academ_sheet.cell(row=6, column=7)
        academ_cell.value = record['academical_tutor']
        academ_cell.alignment = Alignment(horizontal='left')

        wb.save(globals.DATA_FOLDER + 'out/' + record['surname'] + ' ' + record['name'] + '.xlsx')
                
        pbar_rec.update(1)

print("Finished execution.");